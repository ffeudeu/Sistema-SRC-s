import os
import shutil
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import win32com.client as win32
import traceback

# Configurações originais
CAMINHO_SHAREPOINT_LOCAL = r"C:\Users\BR0177332417\OneDrive - Enel Spa\Documents\Nova pasta\Enel Spa\Regularização de Clandestino - Script Diário"

#EMAIL_DESTINO = ["raquel.fernandes@enel.com", "raquel.barros@applus.com", "carla.silva@applus.com", "danilo.mendes@applus.com", "loana.silva@applus.com", "gilmar.santiago@applus.com", "lroccacr@emeal.nttdata.com"] 
#EMAIL_CC = ["leandro.vasconcellos@enel.com", "fabio.casanova@enel.com", "raquel.fernandes@enel.com"]

colunas_finais = [
    'numero_cliente', 'nome', 'numero_ordem', 'cod_doc', 'desc_docto',
    'nro_docto', 'dv_doc', 'municipio', 'bairro', 'data_ingresso',
    'data_estado_ordem', 'tipo_ordem', 'tipo_servico', 'estado_ordem',
    'cod_retorno', 'descricao_retorno', 'tipo_ordemf', 'tipo_servicof',
    'numero_ordem_filha', 'data_ingresso1', 'data_execucao1', 'data_finalizacao1',
    'estado_ordem_filha', 'novo_numero_cliente', 'cod_retornof', 'descricao_retornof',
    'POLO'
]

def processar_arquivo_txt(caminho_arquivo):
    # (MANTIVE SUA FUNÇÃO EXATAMENTE IGUAL)
    dados_processados = []
    with open(caminho_arquivo, 'r', encoding='latin1') as f:
        for linha in f:
            linha = linha.strip()
            if not linha: continue
            partes = [p.strip() for p in linha.split('|')]
            linha_dict = {}
            for i in range(len(colunas_finais) - 1):
                if i < len(partes): linha_dict[colunas_finais[i]] = partes[i]
                else: linha_dict[colunas_finais[i]] = ""

            municipio = linha_dict.get('municipio', '').upper()
            if municipio in ["ANGRA DOS REIS", "BOCAINA MINAS", "ITATIAIA", "MANGARATIBA", "PARATY", "PORTO REAL", "RESENDE"]: linha_dict['POLO'] = "SUL"
            elif municipio in ["CAMPOS DOS GOYTACAZES", "CARDOSO MOREIRA", "S F ITABAPOANA", "S JOAO DA BARRA"]: linha_dict['POLO'] = "CAMPOS"
            elif municipio in ["ARARUAMA", "ARMACAO DOS BUZIOS", "ARRAIAL DO CABO", "CABO FRIO", "IGUABA GRANDE", "SAO PEDRO DA ALDEIA", "SAQUAREMA", "SILVA JARDIM"]: linha_dict['POLO'] = "LAGOS"
            elif municipio in ["CARAPEBUS", "CASIMIRO ABREU", "CONCEI MACABU", "MACAE", "QUISSAMA", "RIO DAS OSTRAS"]: linha_dict['POLO'] = "MACAE"
            elif municipio in ["CACH MACACU", "DUQUE DE CAXIAS", "GUAPIMIRIM", "MAGE"]: linha_dict['POLO'] = "MAGÉ"
            elif municipio in ["MARICA", "NITEROI"]: linha_dict['POLO'] = "NITEROI"
            elif municipio in ["APERIBE", "B J ITABAPOANA", "BOM JARDIM", "CAMBUCI", "CANTAGALO", "CARMO", "CORDEIRO", "DUAS BARRAS", "ITALVA", "ITAOCARA", "ITAPERUNA", "LAJE DO MURIAE", "MACUCO", "MIRACEMA", "NATIVIDADE", "PORCIUNCULA", "S SEBAS DO ALTO", "SAO FIDELIS", "SAO JOSE DE UBA", "STA MA MADALENA", "STO ANTONIO DE PADUA", "TRAJANO DE MORAIS", "TRAJANO DE MORAES", "VARRE-SAI"]: linha_dict['POLO'] = "NOROESTE"
            elif municipio in ["ITABORAI", "RIO BONITO", "SAO GONCALO", "TANGUA"]: linha_dict['POLO'] = "SÃO GONÇALO"
            elif municipio in ["AREAL", "NOVA FRIBURGO", "PARAIBA DO SUL", "PETROPOLIS", "S JOSE VR PRETO", "SUMIDOURO", "TERESOPOLIS", "TRES RIOS"]: linha_dict['POLO'] = "SERRANA"
            else: linha_dict['POLO'] = "--" if municipio else ""
            dados_processados.append(linha_dict)
    return pd.DataFrame(dados_processados)

def gerar_excel(df_consolidado, caminho_saida, fn_progresso):
    # (MANTIVE SUA FUNÇÃO IGUAL, USANDO A CALLBACK DE PROGRESSO)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ordens Consolidadas"
    ws.append(colunas_finais)
    total_linhas = len(df_consolidado)
    for idx, (_, linha) in enumerate(df_consolidado.iterrows()):
        ws.append([linha[col] for col in colunas_finais])
        if idx % max(1, total_linhas // 10) == 0:
            pct = 15 + int((idx / total_linhas) * 30)
            fn_progresso(pct, f"Escrevendo linhas no Excel ({idx}/{total_linhas})...")

    fonte_padrao = 'Segoe UI'
    header_font = Font(name=fonte_padrao, size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    zebra_fill = PatternFill(start_color='F2F5F8', end_color='F2F5F8', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')

    colunas_centralizadas = {'numero_cliente', 'numero_ordem', 'cod_doc', 'desc_docto', 'nro_docto', 'dv_doc', 'data_ingresso', 'data_estado_ordem', 'tipo_ordem', 'tipo_servico', 'cod_retorno', 'tipo_ordemf', 'tipo_servicof', 'numero_ordem_filha', 'data_ingresso1', 'data_execucao1', 'data_finalizacao1', 'novo_numero_cliente', 'cod_retornof', 'POLO'}

    for col_idx in range(1, len(colunas_finais) + 1):
        celula = ws.cell(row=1, column=col_idx)
        celula.font = header_font
        celula.fill = header_fill
        celula.alignment = align_center

    max_linhas_ws = ws.max_row
    for row_idx in range(2, max_linhas_ws + 1):
        cor_linha = zebra_fill if row_idx % 2 == 0 else white_fill
        for col_idx in range(1, len(colunas_finais) + 1):
            celula = ws.cell(row=row_idx, column=col_idx)
            col_nome = colunas_finais[col_idx - 1]
            celula.fill = cor_linha
            celula.border = thin_border
            celula.font = Font(name=fonte_padrao, size=10)
            celula.alignment = align_center if col_nome in colunas_centralizadas else align_left
        if row_idx % max(1, max_linhas_ws // 10) == 0:
            pct = 45 + int((row_idx / max_linhas_ws) * 35)
            fn_progresso(pct, f"Formatando células ({row_idx}/{max_linhas_ws})...")

    fn_progresso(80, "Ajustando largura das colunas...")
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 13)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(colunas_finais))}{ws.max_row}"
    ws.freeze_panes = "A2"
    fn_progresso(85, "Salvando arquivo Excel...")
    wb.save(caminho_saida)

def enviar_email(caminho_arquivo, emails_to, emails_cc):
    try:
        try: outlook = win32.GetActiveObject("Outlook.Application")
        except: outlook = win32.DispatchEx('Outlook.Application')
            
        mail = outlook.CreateItem(0)
        mail.To = "; ".join(emails_to)
        if emails_cc: mail.CC = "; ".join(emails_cc)
        mail.Subject = "Script Ordem Gerada"
        url_link = "https://enelcom.sharepoint.com/:f:/r/sites/clandestine_connection_solution/Documentos%20Compartilhados/Script%20Di%C3%A1rio?csf=1&web=1&e=otLdSv"
        
        mail.HTMLBody = f"""
        <html>
            <body style="font-family: Arial, sans-serif; font-size: 14px;">
                <p>Olá,</p>
                <p>Segue em anexo o Script atualizado.</p>
                <p><a href="{url_link}" style="color: #1F4E78; font-weight: bold;">Script Diário</a></p>
                <br><p><b>Data/Hora:</b><br>{datetime.now().strftime('%d/%m/%Y %H:%M')}</p>
                <br><p><b>João Lucas L. P. Antunes</b><br>Clandestine Connections Management<br>Customer Engagement Rio<br>I&N Rio<br></p>
            </body>
        </html>
        """
        mail.Attachments.Add(caminho_arquivo)
        mail.Send()
        return True, None
    except Exception as e:
        return False, traceback.format_exc()

# ==========================================
# FUNÇÃO PRINCIPAL (MOTOR) CHAMADA PELA INTERFACE
# ==========================================
def executar_automacao_backend(arquivo1, arquivo2, pasta_saida, callback_progresso, emails_to, emails_cc):
    """
    Esta função roda todo o processo. Em vez de mexer na tela, ela chama 
    a `callback_progresso(porcentagem, texto)` para avisar a tela do que está acontecendo.
    Retorna uma string com a mensagem de sucesso para a tela exibir.
    """
    callback_progresso(5, "Lendo arquivo: Ordem Gerada...")
    df_gerada = processar_arquivo_txt(arquivo1)
    
    callback_progresso(10, "Lendo arquivo: Ordem VIC...")
    df_vic = processar_arquivo_txt(arquivo2)

    callback_progresso(15, "Consolidando dados...")
    df_consolidado = pd.concat([df_gerada, df_vic], ignore_index=True)

    nome_arquivo = f"Consolidado_Ordens_Rio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    caminho_saida = os.path.join(pasta_saida, nome_arquivo)

    gerar_excel(df_consolidado, caminho_saida, callback_progresso)

    sp_atualizado = False
    if CAMINHO_SHAREPOINT_LOCAL and os.path.exists(CAMINHO_SHAREPOINT_LOCAL):
        callback_progresso(88, "Limpando arquivos antigos do SharePoint...")
        for arquivo_pasta in os.listdir(CAMINHO_SHAREPOINT_LOCAL):
            if arquivo_pasta.startswith("Consolidado_Ordens_Rio") and arquivo_pasta.endswith(".xlsx"):
                try: os.remove(os.path.join(CAMINHO_SHAREPOINT_LOCAL, arquivo_pasta))
                except: pass
        
        callback_progresso(92, "Copiando novo arquivo para o SharePoint...")
        shutil.copy2(caminho_saida, os.path.join(CAMINHO_SHAREPOINT_LOCAL, nome_arquivo))
        sp_atualizado = True

    callback_progresso(95, "Conectando ao servidor e enviando e-mail...")
    sucesso_email, erro_email = enviar_email(caminho_saida, emails_to, emails_cc)
    # ========================================
    # ENVIO DE E-MAIL 
    # ========================================
    
    if not emails_to and not emails_cc:
        callback_progresso(100, "Arquivo Excel gerado com sucesso!")
        return caminho_saida
        
    # Se as listas tiverem conteúdo (versão .exe), envia o e-mail pelo Outlook normalmente
    else:
        callback_progresso(95, "Abrindo o Outlook para envio de e-mail...")
        sucesso_email, erro_email = enviar_email(caminho_saida, emails_to, emails_cc)
        
        callback_progresso(100, "Processo finalizado!")
        if sucesso_email:
            return "Unificação concluída e e-mail gerado com sucesso no Outlook!"
        else:
            return f"Unificação concluída, mas falhou ao gerar o e-mail.\nErro: {erro_email}"